import re
import time
import openpyxl
from deep_translator import GoogleTranslator

INPUT_FILE = "C:/Users/alonr/OneDrive/Documents/LLM creativity/Experiment 1/Users_keepable_paired_only_corrected_audit.xlsx"
OUTPUT_FILE = "C:/Users/alonr/OneDrive/Documents/LLM creativity/Experiment 1/Users_keepable_paired_only_corrected_audit_translated.xlsx"
SHEET_NAME = "corrected_users"

HEBREW_PATTERN = re.compile(r'[א-ת]')

def translate(text: str) -> str:
    for attempt in range(5):
        try:
            translator = GoogleTranslator(source="auto", target="en")
            result = translator.translate(str(text))
            time.sleep(0.5)  # polite delay between requests
            return result
        except Exception as e:
            wait = 10 * (attempt + 1)
            print(f"    Retry {attempt+1}/5 after {wait}s ({e})")
            time.sleep(wait)
    raise RuntimeError(f"Failed to translate after 5 attempts: {text[:50]}")

def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb[SHEET_NAME]

    headers = [cell.value for cell in ws[1]]
    total_translated = 0

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value and HEBREW_PATTERN.search(str(cell.value)):
                col_name = headers[cell.column - 1]
                print(f"  Row {cell.row}, col '{col_name}': translating...")
                cell.value = translate(cell.value)
                total_translated += 1

    wb.save(OUTPUT_FILE)
    print(f"\nDone. Translated {total_translated} cells. Output: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
