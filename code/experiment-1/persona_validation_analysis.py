"""
Persona validation analysis.
Two complementary analyses:
  1. Argmax hit-rate: did the user give the highest score to their persona's question?
  2. Mean-score discrimination: is the matched-question mean significantly higher than others?
Produces a 3-panel figure (confusion matrix, mean scores, hit rates).
"""
import openpyxl
import numpy as np
import pandas as pd
from scipy import stats
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import warnings, pathlib

warnings.filterwarnings("ignore")

HERE = pathlib.Path(r"C:\Users\alonr\OneDrive\Documents\LLM creativity\Experiment 1")
XLSX = pathlib.Path(r"C:\Users\alonr\AppData\Local\Temp\questionnaire.xlsx")
OUT  = HERE / "figures"
OUT.mkdir(exist_ok=True)

# ── load ──────────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX)
ws = wb["corrected_users"]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
data = pd.DataFrame(list(ws.iter_rows(min_row=2, values_only=True)), columns=headers)

# ── column mapping ─────────────────────────────────────────────────────────────
VAL_COLS = {
    "BoundedRational": "Bounded-Rational validation  ",
    "Rational":        "Rational Validation ",
    "Divergent":       "Divergent Thinking Validation ",
    "Convergent":      "Convergent thinking validation",
}
PERSONA_NORM = {
    "bounded rationality": "BoundedRational",
    "strictly rational":   "Rational",
    "divergent":           "Divergent",
    "convergent":          "Convergent",
}
LABELS = ["BoundedRational", "Rational", "Divergent", "Convergent"]
PRETTY = {
    "BoundedRational": "Bounded\nRational",
    "Rational":        "Rational",
    "Divergent":       "Divergent",
    "Convergent":      "Convergent",
}
COLORS = {
    "BoundedRational": "#4472C4",
    "Rational":        "#ED7D31",
    "Divergent":       "#70AD47",
    "Convergent":      "#FFC000",
}

# ── clean & filter ─────────────────────────────────────────────────────────────
df = data[["Persona"] + list(VAL_COLS.values())].copy()
df.columns = ["persona"] + LABELS
df["persona"] = df["persona"].str.strip().str.lower().map(PERSONA_NORM)
df = df.dropna(subset=["persona"])
for col in LABELS:
    df[col] = pd.to_numeric(df[col], errors="coerce")
df = df.dropna(subset=LABELS)
df = df[df["persona"].isin(LABELS)]
print(f"Usable rows: {len(df)}  |  per-persona n: {df['persona'].value_counts().to_dict()}")

# ── argmax prediction (first-occurrence tie-breaking = deterministic) ──────────
df["predicted"] = df[LABELS].idxmax(axis=1)
df["hit"]       = (df["persona"] == df["predicted"])

# ── Confusion matrix ───────────────────────────────────────────────────────────
cm = pd.crosstab(df["persona"], df["predicted"]).reindex(
    index=LABELS, columns=LABELS, fill_value=0
)
print("\nConfusion matrix (counts):\n", cm)

# ── Analysis 1: argmax hit-rate + binomial tests ───────────────────────────────
chance = 1 / len(LABELS)
hit_results = []
for persona in LABELS:
    sub  = df[df["persona"] == persona]
    n    = len(sub)
    hits = int(sub["hit"].sum())
    rate = hits / n if n else np.nan
    binom = stats.binomtest(hits, n, p=chance, alternative="greater")
    hit_results.append(dict(persona=persona, n=n, hits=hits,
                             hit_rate=rate, p_binom=binom.pvalue))
hits_df = pd.DataFrame(hit_results)
print("\nHit-rate results:\n", hits_df.to_string(index=False))

hit_total = df["hit"].sum()
n_total   = len(df)
binom_overall = stats.binomtest(int(hit_total), n_total, p=chance, alternative="greater")
chi2_full, p_full, dof_full, _ = stats.chi2_contingency(cm.values)
print(f"\nOverall: {hit_total}/{n_total} = {hit_total/n_total:.1%}")
print(f"Binomial p = {binom_overall.pvalue:.4f}")
print(f"χ²({dof_full}) = {chi2_full:.2f}, p = {p_full:.4f}")

# ── Analysis 2: mean-score discrimination (matched vs mean-of-others) ──────────
mean_results = []
for persona in LABELS:
    sub   = df[df["persona"] == persona].copy()
    n     = len(sub)
    match = sub[persona].values
    other_cols = [c for c in LABELS if c != persona]
    other_mean = sub[other_cols].mean(axis=1).values
    diff  = match - other_mean
    t, p  = stats.ttest_1samp(diff, 0)
    mean_results.append(dict(
        persona=persona, n=n,
        mean_match=float(np.mean(match)),
        mean_other=float(np.mean(other_mean)),
        mean_diff=float(np.mean(diff)),
        t=t, p_ttest=p,
    ))
means_df = pd.DataFrame(mean_results)
print("\nMean-score discrimination:\n", means_df.to_string(index=False))

# ── Figure: 3 panels ───────────────────────────────────────────────────────────
fig = plt.figure(figsize=(17, 6), facecolor="white")
fig.subplots_adjust(wspace=0.42, left=0.06, right=0.97, top=0.87, bottom=0.14)

gs = fig.add_gridspec(1, 3, width_ratios=[1.6, 1.7, 1.3])
ax1 = fig.add_subplot(gs[0])   # confusion matrix
ax2 = fig.add_subplot(gs[1])   # mean scores grouped bar
ax3 = fig.add_subplot(gs[2])   # hit rates

pretty_labels = [PRETTY[l] for l in LABELS]

# ── Panel A: confusion matrix ──────────────────────────────────────────────────
cm_arr = cm.values.astype(float)
cm_pct = cm_arr / cm_arr.sum(axis=1, keepdims=True) * 100

im = ax1.imshow(cm_pct, cmap="Blues", vmin=0, vmax=100, aspect="auto")
for i in range(len(LABELS)):
    for j in range(len(LABELS)):
        count = int(cm_arr[i, j])
        pct   = cm_pct[i, j]
        clr   = "white" if pct > 58 else "black"
        ax1.text(j, i, f"{count}\n({pct:.0f}%)",
                 ha="center", va="center", fontsize=9.5,
                 color=clr, fontweight="bold" if i == j else "normal")
        if i == j:
            ax1.add_patch(plt.Rectangle(
                (j - 0.5, i - 0.5), 1, 1,
                fill=False, edgecolor="#DD2222", lw=2.2))

ax1.set_xticks(range(len(LABELS)))
ax1.set_xticklabels(pretty_labels, fontsize=9.5)
ax1.set_yticks(range(len(LABELS)))
ax1.set_yticklabels(pretty_labels, fontsize=9.5)
ax1.set_xlabel("Highest-rated validation question", fontsize=10, labelpad=7)
ax1.set_ylabel("Assigned persona", fontsize=10, labelpad=7)
ax1.set_title("A. Confusion matrix\n(row-normalised %)", fontsize=11, pad=8)
cbar = fig.colorbar(im, ax=ax1, shrink=0.80, pad=0.02)
cbar.ax.yaxis.set_major_formatter(mticker.FormatStrFormatter("%d%%"))
cbar.ax.tick_params(labelsize=8)

# ── Panel B: mean scores per persona group (grouped bar) ──────────────────────
x = np.arange(len(LABELS))
bar_w = 0.19
offsets = np.linspace(-(len(LABELS)-1)/2 * bar_w, (len(LABELS)-1)/2 * bar_w, len(LABELS))

for bi, q_persona in enumerate(LABELS):
    means = []
    for assigned in LABELS:
        sub = df[df["persona"] == assigned]
        means.append(sub[q_persona].mean())
    bars = ax2.bar(x + offsets[bi], means, width=bar_w,
                   color=COLORS[q_persona], alpha=0.88,
                   label=PRETTY[q_persona].replace("\n", " "), zorder=3)

# bold outline on diagonal bars (matched question for each persona)
for bi, q_persona in enumerate(LABELS):
    ai = LABELS.index(q_persona)    # assigned persona = same
    sub = df[df["persona"] == q_persona]
    m = sub[q_persona].mean()
    ax2.bar(x[ai] + offsets[bi], m, width=bar_w,
            color=COLORS[q_persona], edgecolor="black", linewidth=1.8,
            zorder=4, label=None)

ax2.axhline(3.0, color="grey", ls=":", lw=1.1, alpha=0.6)
ax2.set_ylim(1, 5.4)
ax2.set_yticks([1, 2, 3, 4, 5])
ax2.set_ylabel("Mean validation score (1–5)", fontsize=10)
ax2.set_xticks(x)
ax2.set_xticklabels(pretty_labels, fontsize=9.5)
ax2.set_xlabel("Assigned persona group", fontsize=10, labelpad=7)
ax2.set_title("B. Mean scores per persona group\n(bold outline = matched question)", fontsize=11, pad=8)
ax2.yaxis.grid(True, alpha=0.3, zorder=1)
ax2.set_axisbelow(True)

# t-test stars above matched bar
for i, row in means_df.iterrows():
    p = row["p_ttest"]
    ai = LABELS.index(row["persona"])
    bi = LABELS.index(row["persona"])
    m  = row["mean_match"]
    stars = "***" if p < 0.001 else "**" if p < 0.01 else "*" if p < 0.05 else "ns"
    c = "#CC0000" if stars != "ns" else "#666666"
    ax2.text(x[ai] + offsets[bi], m + 0.12, stars,
             ha="center", va="bottom", fontsize=10, color=c, fontweight="bold")

ax2.legend(title="Validation\nquestion", fontsize=8, title_fontsize=8,
           loc="upper right", ncol=2, framealpha=0.9)

# ── Panel C: hit rates + binomial significance ─────────────────────────────────
xs = np.arange(len(LABELS))
bar_colors = [COLORS[p] for p in LABELS]
ax3.bar(xs, hits_df["hit_rate"] * 100, color=bar_colors,
        width=0.55, edgecolor="white", linewidth=0.8, zorder=3)
ax3.axhline(25, color="grey", ls="--", lw=1.4, label="Chance (25%)", zorder=2)
ax3.set_ylim(0, 100)
ax3.set_ylabel("Hit rate (%)", fontsize=10)
ax3.set_title("C. Argmax hit rate per persona\n(vs 25% chance)", fontsize=11, pad=8)
ax3.set_xticks(xs)
ax3.set_xticklabels(pretty_labels, fontsize=9.5)
ax3.yaxis.grid(True, alpha=0.35, zorder=1)
ax3.set_axisbelow(True)

for i, row in hits_df.iterrows():
    p, rate, n, hits = row["p_binom"], row["hit_rate"], row["n"], row["hits"]
    stars = "***" if p < 0.001 else "**" if p < 0.01 else "*" if p < 0.05 else "ns"
    c = "#CC0000" if stars != "ns" else "#666666"
    ax3.text(i, rate * 100 + 2.5, stars,
             ha="center", va="bottom", fontsize=12, color=c, fontweight="bold")
    ax3.text(i, -6.5, f"n={n}\n{hits}/{n}", ha="center", va="top", fontsize=8.5, color="#444")

ax3.text(
    0.97, 0.97,
    f"Total: {hit_total}/{n_total} ({hit_total/n_total:.0%})\n"
    f"Binomial p={binom_overall.pvalue:.3f}\n"
    f"χ²({dof_full})={chi2_full:.1f}, p={p_full:.3f}",
    transform=ax3.transAxes, ha="right", va="top", fontsize=8.2,
    bbox=dict(boxstyle="round,pad=0.4", fc="lightyellow", ec="#CCAA00", lw=1),
)
ax3.legend(fontsize=9, loc="upper left")

fig.suptitle(
    "Persona Validation Check: Did users score their assigned persona's question highest?",
    fontsize=13, fontweight="bold", y=0.97,
)

out_path = OUT / "fig_persona_validation_confusion.png"
fig.savefig(out_path, dpi=160, bbox_inches="tight")
plt.close()
print(f"\nFigure saved: {out_path}")
